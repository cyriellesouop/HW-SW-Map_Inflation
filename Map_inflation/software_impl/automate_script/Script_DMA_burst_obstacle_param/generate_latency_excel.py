#!/usr/bin/env python3
"""
generate_latency_excel.py

Runs the fpga_cycles binary for a list of obstacle counts and writes
results to an Excel workbook. Each (kernel_size, n_obstacles) pair
gets its own sheet named: latency_result_{kernel_size}_{n_obstacles}
A summary sheet collects all rows for easy comparison.

Usage:
    python generate_latency_excel.py \
        --binary ./fpga_cycles \
        --cpu_bram 3 \
        --dma_ip 4 \
        --latency 25 \
        --kernel 3 \
        --obstacles 100 500 1000 5000 9810
"""

import argparse
import subprocess
import sys
import io
import csv
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------
# Styling helpers
# ---------------------------------------------------------------
HEADER_FILL   = PatternFill("solid", start_color="1F4E79", end_color="1F4E79")
HEADER_FONT   = Font(name="Arial", bold=True, color="FFFFFF", size=11)
TITLE_FONT    = Font(name="Arial", bold=True, size=13)
CELL_FONT     = Font(name="Arial", size=10)
LABEL_FONT    = Font(name="Arial", bold=True, size=10)
STAGE_FILL    = PatternFill("solid", start_color="D6E4F0", end_color="D6E4F0")
TOTAL_FILL    = PatternFill("solid", start_color="FCE4D6", end_color="FCE4D6")
TOTAL_FONT    = Font(name="Arial", bold=True, size=10)
CENTER        = Alignment(horizontal="center", vertical="center")
LEFT          = Alignment(horizontal="left",   vertical="center")

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

COLS = [
    "kernel_size", "burst_length", "n_obstacles", "n_values", "n_transfers",
    "n_full_bursts", "n_partial_burst_words",
    "cycles_dma", "cycles_dma_to_ip", "cycles_ip_to_systolic",
    "cycles_systolic", "cycles_total"
]

HEADERS = [
    "Kernel Size", "Burst Length", "Obstacles", "Total Values", "AXI Transfers",
    "Full Bursts", "Partial Burst Words",
    "DMA Cycles", "DMA→IP Cycles", "IP→Systolic Cycles",
    "Systolic Cycles", "Total Cycles"
]

COL_WIDTHS = [14, 14, 13, 15, 16, 14, 22, 14, 16, 20, 17, 14]


def style_header_row(ws, row, ncols):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = HEADER_FONT
        cell.fill      = HEADER_FILL
        cell.alignment = CENTER
        cell.border    = THIN_BORDER


def style_data_row(ws, row, ncols, fill=None):
    for col in range(1, ncols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font      = TOTAL_FONT if fill == TOTAL_FILL else CELL_FONT
        cell.alignment = CENTER
        cell.border    = THIN_BORDER
        if fill:
            cell.fill = fill


def set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ---------------------------------------------------------------
# Build one detail sheet per (kernel_size, n_obstacles)
# ---------------------------------------------------------------
def build_detail_sheet(wb, row_data, params):
    k    = row_data["kernel_size"]
    obs  = row_data["n_obstacles"]
    bl   = row_data["burst_length"]
    sheet_name = f"latency_result_{k}_{obs}"
    ws = wb.create_sheet(title=sheet_name)

    # --- Title ---
    ws.merge_cells("A1:D1")
    ws["A1"] = f"Latency Breakdown — Kernel {k}×{k}, {obs:,} Obstacles, Burst {bl}"
    ws["A1"].font      = TITLE_FONT
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 22

    # --- Parameters block ---
    param_labels = [
        ("Cycles CPU → BRAM (per transfer)",        params["cpu_bram"]),
        ("DMA Burst Length (beats)",                 bl),
        ("AXI Data Width (bits)",                    32),
        ("Pixel Width (bits)",                        8),
        ("Values per Beat",                           4),
        ("Cycles DMA → IP input",                   params["dma_ip"]),
        ("Cycles IP input → Systolic array input",  params["ip_systolic"]),
        ("Systolic Pipeline Latency (cycles)",       params["latency"]),
    ]
    ws["A3"] = "System Parameters"
    ws["A3"].font = LABEL_FONT
    for i, (label, val) in enumerate(param_labels, start=4):
        ws.cell(row=i, column=1).value     = label
        ws.cell(row=i, column=1).font      = CELL_FONT
        ws.cell(row=i, column=1).alignment = LEFT
        ws.cell(row=i, column=2).value     = val
        ws.cell(row=i, column=2).font      = LABEL_FONT
        ws.cell(row=i, column=2).alignment = CENTER

    # --- Data table ---
    tbl_row = 4 + len(param_labels) + 2

    fields = [
        ("Input Data",         None),
        ("  Total values",              row_data["n_values"],                 None),
        ("  AXI transfers",             row_data["n_transfers"],              None),
        ("  Full bursts",               row_data["n_full_bursts"],            None),
        ("  Partial burst",             row_data["n_partial_burst_words"],    None),
        ("Pipeline Stages",    None),
        ("  Stage 1 — DMA burst transfer",        row_data["cycles_dma"],             STAGE_FILL),
        ("  Stage 2 — DMA → IP input latency",    row_data["cycles_dma_to_ip"],       STAGE_FILL),
        ("  Stage 3 — IP input → Systolic input", row_data["cycles_ip_to_systolic"],  STAGE_FILL),
        ("  Stage 4 — Systolic array",            row_data["cycles_systolic"],        STAGE_FILL),
        ("Total Hardware Cycles",                 row_data["cycles_total"],           TOTAL_FILL),
    ]

    ws.cell(row=tbl_row, column=1).value = "Description"
    ws.cell(row=tbl_row, column=2).value = "Value"
    ws.cell(row=tbl_row, column=1).font  = HEADER_FONT
    ws.cell(row=tbl_row, column=2).font  = HEADER_FONT
    ws.cell(row=tbl_row, column=1).fill  = HEADER_FILL
    ws.cell(row=tbl_row, column=2).fill  = HEADER_FILL
    ws.cell(row=tbl_row, column=1).alignment = CENTER
    ws.cell(row=tbl_row, column=2).alignment = CENTER

    for i, (label, val, *rest) in enumerate(fields, start=tbl_row + 1):
        fill = rest[0] if rest else None
        c1 = ws.cell(row=i, column=1, value=label)
        c1.font      = TOTAL_FONT if fill == TOTAL_FILL else CELL_FONT
        c1.alignment = LEFT
        c1.border    = THIN_BORDER
        if fill:
            c1.fill = fill
        if val is not None:
            c2 = ws.cell(row=i, column=2, value=val)
            c2.font      = TOTAL_FONT if fill == TOTAL_FILL else CELL_FONT
            c2.alignment = CENTER
            c2.border    = THIN_BORDER
            if fill:
                c2.fill = fill

    ws.column_dimensions["A"].width = 38
    ws.column_dimensions["B"].width = 18


# ---------------------------------------------------------------
# Build summary sheet
# ---------------------------------------------------------------
def build_summary_sheet(wb, all_rows):
    ws = wb.active
    ws.title = "Summary"

    ws.merge_cells(f"A1:{get_column_letter(len(HEADERS))}1")
    ws["A1"] = "FPGA Inflation Latency — Summary"
    ws["A1"].font      = TITLE_FONT
    ws["A1"].alignment = LEFT
    ws.row_dimensions[1].height = 22

    # header row
    for col, h in enumerate(HEADERS, start=1):
        cell = ws.cell(row=2, column=col, value=h)
    style_header_row(ws, 2, len(HEADERS))

    # data rows
    for r, row in enumerate(all_rows, start=3):
        fill = TOTAL_FILL if r == len(all_rows) + 2 else None
        for col, key in enumerate(COLS, start=1):
            ws.cell(row=r, column=col, value=row[key])
        style_data_row(ws, r, len(HEADERS), fill=fill)

    set_col_widths(ws, COL_WIDTHS)
    ws.freeze_panes = "A3"


# ---------------------------------------------------------------
# Parse CSV output from the C binary
# ---------------------------------------------------------------
def run_binary(args, obstacles):
    cmd = [
        args.binary,
        str(args.cpu_bram),
        str(args.dma_ip),
        str(args.ip_systolic),
        str(args.latency),
        str(args.kernel),
        str(args.burst_length),
    ] + [str(o) for o in obstacles]

    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        print(f"Error running binary:\n{result.stderr}", file=sys.stderr)
        sys.exit(1)

    rows = []
    reader = csv.DictReader(io.StringIO(result.stdout))
    for row in reader:
        rows.append({k: int(v) for k, v in row.items()})
    return rows


# ---------------------------------------------------------------
# Main
# ---------------------------------------------------------------
def main():
    parser = argparse.ArgumentParser(
        description="Run fpga_cycles and write latency results to Excel."
    )
    parser.add_argument("--binary",       default="./fpga_cycles",
                        help="Path to the compiled fpga_cycles binary")
    parser.add_argument("--cpu_bram",     type=int, required=True,
                        help="Clock cycles per CPU→BRAM transfer")
    parser.add_argument("--dma_ip",       type=int, required=True,
                        help="Clock cycles from DMA to IP input interface")
    parser.add_argument("--ip_systolic",  type=int, required=True,
                        help="Clock cycles from IP input ports to systolic array input")
    parser.add_argument("--latency",      type=int, required=True,
                        help="Systolic array pipeline latency (cycles)")
    parser.add_argument("--kernel",       type=int, default=3,
                        help="Kernel size K (default: 3)")
    parser.add_argument("--burst_length", type=int, default=16,
                        help="DMA burst length in beats (default: 16)")
    parser.add_argument("--obstacles",    type=int, nargs="+", required=True,
                        help="List of obstacle counts to evaluate")
    args = parser.parse_args()

    # run C binary and collect rows
    all_rows = run_binary(args, args.obstacles)

    # build workbook
    wb = Workbook()

    # summary sheet (uses the default active sheet)
    build_summary_sheet(wb, all_rows)

    # one detail sheet per row
    for row in all_rows:
        build_detail_sheet(wb, row, {
            "cpu_bram":     args.cpu_bram,
            "dma_ip":       args.dma_ip,
            "ip_systolic":  args.ip_systolic,
            "latency":      args.latency,
            "burst_length": args.burst_length,
        })

    # save — filename includes kernel size, burst length, and obstacle list
    k   = args.kernel
    bl  = args.burst_length
    obs = "_".join(str(o) for o in args.obstacles)
    filename = f"latency_results_k{k}_burst{bl}.xlsx"
    wb.save(filename)
    print(f"Saved: {filename}")
    print(f"  Sheets: Summary + {len(all_rows)} detail sheets")


if __name__ == "__main__":
    main()
